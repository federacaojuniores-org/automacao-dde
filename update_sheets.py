import os
import openpyxl
from datetime import datetime, timedelta
from google.oauth2 import service_account
from googleapiclient.discovery import build
from dotenv import load_dotenv

# Load local environment variables from .env file
load_dotenv()

# Dynamically resolve project directory
PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))

# Get credentials file path (resolve relative path to project directory)
credentials_filename = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "dde-projeto-9ed22179e048.json")
if os.path.isabs(credentials_filename):
    CREDENTIALS_PATH = credentials_filename
else:
    CREDENTIALS_PATH = os.path.join(PROJECT_DIR, credentials_filename)

SPREADSHEET_ID = os.getenv("GOOGLE_SPREADSHEET_ID", "163X5ADTJkHXK4INVs4KPdAXveUXhz0sYEoDGIdHWdOM")
DOWNLOAD_DIR = os.path.join(PROJECT_DIR, "downloads")

REPORTS_TO_UPDATE = [
    {
        "filename": "monitoramento_geral_v2.xlsx",
        "sheet_name": "[MONITORAMENTO] Geral v2.0"
    },
    {
        "filename": "monitoramento_acumulado_v2.xlsx",
        "sheet_name": "[MONITORAMENTO] Acumulado v2.0"
    },
    {
        "filename": "empresas_juniores_geral_v2.xlsx",
        "sheet_name": "[EMPRESAS JUNIORES] Geral v2.0"
    }
]

def format_value(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return val

def update_sheet_auto_aligned(service, filename, sheet_name):
    excel_path = os.path.join(DOWNLOAD_DIR, filename)
    print(f"\n--- Reading Excel file: {filename} ---")
    
    # 1. Load Excel headers and data in streaming read_only mode
    wb_excel = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet_excel = wb_excel.active
    
    # Get Excel headers (first row)
    excel_rows = sheet_excel.iter_rows(values_only=True)
    try:
        excel_headers = [str(h).strip().upper() for h in next(excel_rows) if h is not None]
    except StopIteration:
        print("  [Error] Excel file is empty!")
        wb_excel.close()
        return
        
    print(f"  Excel active sheet: '{sheet_excel.title}'")
    print(f"  Found {len(excel_headers)} headers in Excel.")
    
    # Read all remaining data rows from Excel
    excel_data = []
    for row in excel_rows:
        excel_data.append([format_value(val) for val in row])
    wb_excel.close()
    print(f"  Loaded {len(excel_data)} data rows from Excel.")

    # 2. Load ORIGINAL headers from the restored Google Sheet
    print(f"  Reading original headers from Google Sheet tab: '{sheet_name}'...")
    sheet_metadata = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{sheet_name}'!A1:ZZ1"
    ).execute()
    
    gs_headers_raw = sheet_metadata.get("values", [])[0]
    gs_headers = [str(h).strip() for h in gs_headers_raw]
    print(f"  Found {len(gs_headers)} original headers in Google Sheet.")

    # Find new Excel headers that are not in the Google Sheet (case-insensitive check)
    gs_headers_upper = {h.strip().upper() for h in gs_headers}
    # Read raw Excel headers to preserve their original casing
    wb_excel_raw = openpyxl.load_workbook(excel_path, read_only=True)
    sheet_excel_raw = wb_excel_raw.active
    raw_headers_rows = sheet_excel_raw.iter_rows(values_only=True)
    raw_excel_headers = [str(h).strip() for h in next(raw_headers_rows) if h is not None]
    wb_excel_raw.close()

    new_headers = []
    for raw_h in raw_excel_headers:
        if raw_h.strip().upper() not in gs_headers_upper:
            new_headers.append(raw_h)
            
    if new_headers:
        print(f"  [New Columns Detected] Appending {len(new_headers)} new columns at the end of the sheet: {new_headers}")
        gs_headers.extend(new_headers)

    # 3. Build the aligned upload matrix
    aligned_values = [gs_headers] # First row is always the original Google Sheet headers (NEVER renamed!)
    
    # Pre-map Excel header positions for O(1) lookups
    excel_header_map = {name: idx for idx, name in enumerate(excel_headers)}
    
    print("  Aligning Excel columns with Google Sheet structure by name...")
    for row_idx, excel_row in enumerate(excel_data):
        aligned_row = []
        for gs_h in gs_headers:
            gs_h_upper = gs_h.strip().upper()
            if gs_h_upper in excel_header_map:
                excel_idx = excel_header_map[gs_h_upper]
                # Extract value from Excel row at that index
                if excel_idx < len(excel_row):
                    aligned_row.append(excel_row[excel_idx])
                else:
                    aligned_row.append("")
            else:
                # If column doesn't exist in Excel, write empty (allows Table formula propagation)
                aligned_row.append("")
        aligned_values.append(aligned_row)

    # 4. Clear old data from Google Sheets (preserving headers and Table structure)
    print(f"  Clearing old values in Google Sheet tab: '{sheet_name}' (preserving headers)...")
    service.spreadsheets().values().clear(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{sheet_name}'!A1:ZZ100000"
    ).execute()
    
    # 5. Write the aligned values starting at A1 in chunks (robust against socket timeouts!)
    chunk_size = 2000
    print(f"  Writing aligned values in chunks of {chunk_size} rows in tab: '{sheet_name}'...")
    
    for i in range(0, len(aligned_values), chunk_size):
        chunk = aligned_values[i : i + chunk_size]
        start_row = i + 1
        end_row = i + len(chunk)
        print(f"    Writing rows {start_row} to {end_row}...")
        service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=f"'{sheet_name}'!A{start_row}",
            valueInputOption="USER_ENTERED",
            body={"values": chunk}
        ).execute()
    
    print(f"  SUCCESS: Updated {len(aligned_values)} rows in Google Sheets!")

def main():
    print("Authenticating with Google Sheets API...")
    creds = service_account.Credentials.from_service_account_file(
        CREDENTIALS_PATH,
        scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    service = build("sheets", "v4", credentials=creds)
    print("Authentication complete.")
    
    for report in REPORTS_TO_UPDATE:
        update_sheet_auto_aligned(service, report["filename"], report["sheet_name"])
        
    # Final Step: Automatically update the Dashboard reference date in C3 to Brasília today's date!
    now_brasilia = datetime.utcnow() - timedelta(hours=3)
    today_str = now_brasilia.strftime("%d/%m/%Y")
    print(f"\n--- Updating reference date in '[REDE] Dashboard'!C3 to Brasília date: {today_str} ---")
    service.spreadsheets().values().update(
        spreadsheetId=SPREADSHEET_ID,
        range="'[REDE] Dashboard'!C3",
        valueInputOption="USER_ENTERED",
        body={"values": [[today_str]]}
    ).execute()
    print("Dashboard reference date updated successfully!")
    
    # Update the last updated log in B4 (merged B-H4) to: "Dados atualizados pela última vez em: dd/MM/yy - hh:mm" (Brasília Time!)
    now_str = now_brasilia.strftime("%d/%m/%y - %H:%M")
    timestamp_text = f"Dados atualizados pela última vez em: {now_str}"
    print(f"--- Updating last updated log in '[REDE] Dashboard'!B4 to: '{timestamp_text}' ---")
    service.spreadsheets().values().update(
        spreadsheetId=SPREADSHEET_ID,
        range="'[REDE] Dashboard'!B4",
        valueInputOption="USER_ENTERED",
        body={"values": [[timestamp_text]]}
    ).execute()
    print("Dashboard last updated log updated successfully!")
    
    print("\nAll 3 destination sheets and the Dashboard reference date have been successfully updated with self-aligned columns!")

if __name__ == "__main__":
    main()
